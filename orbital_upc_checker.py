# --------------------------  UPC‑coverage checker  --------------------------
# ❶  Place this script in the same folder as your seven Orbital spreadsheets.
# ❷  Run:  python orbital_upc_checker.py
#
# It will create three CSV files:
#   • orbital_upc_presence_matrix.csv       – True/False matrix for every file
#   • orbital_upc_missing_breakdown.csv     – UPCs in metadata/products but not in delivery
#   • orbital_upc_summary_counts.csv        – one‑line totals per file
#
# ---------------------------------------------------------------------------

import os, re, csv, zipfile, sys
from pathlib import Path
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def upcs_from_csv(path: Path) -> set[str]:
    """Return a set of UPC strings from the first column whose name contains 'upc'."""
    # detect delimiter from the first line
    with path.open("r", encoding="utf‑8", errors="ignore") as fh:
        sample = fh.readline()
    sep = ";" if sample.count(";") > sample.count(",") else ","

    df = pd.read_csv(path, sep=sep, dtype=str,
                     usecols=lambda c: re.search(r"upc", str(c), re.I))
    if df.empty:
        return set()
    series = df.iloc[:, 0].dropna().str.strip()
    return set(series[series.str.match(r"^\d{8,18}$")])

def upcs_from_xlsx(path: Path, cap_rows: int | None = None) -> set[str]:
    """
    Stream an .xlsx file in read‑only mode and collect UPCs from the first column
    whose header contains 'upc'.  If cap_rows is set, stops after that many data rows.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    upc_idx = None
    upcs: set[str] = set()

    for row_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_i == 1:  # header
            for j, cell in enumerate(row):
                if cell and re.search(r"upc", str(cell), re.I):
                    upc_idx = j
                    break
            if upc_idx is None:
                wb.close()
                return set()
            continue

        if cap_rows and row_i > cap_rows:
            break

        val = row[upc_idx] if upc_idx < len(row) else None
        if val is None:
            continue
        s = str(val).strip()
        if re.fullmatch(r"\d{8,18}", s):
            upcs.add(s)

    wb.close()
    return upcs

# ---------------------------------------------------------------------------
# File list  (add/remove here if the filenames change)
# ---------------------------------------------------------------------------
FILES: dict[str, Path] = {
    "Orbital Track Metadata.csv":                       Path("Orbital Track Metadata.csv"),
    "Orbital Track Metadata 4.csv":                    Path("Orbital Track Metadata 4.csv"),
    "Orbital Track Metadata 3.csv":                    Path("Orbital Track Metadata 3.csv"),
    "Orbital Track Metadata 2.csv":                    Path("Orbital Track Metadata 2.csv"),
    "Orbital Track Metadata (1).csv":                  Path("Orbital Track Metadata (1).csv"),
    "Orbital Media products with territories (1).csv": Path("Orbital Media products with territories (1).csv"),
    "Orbital_Delivery_Schedule.csv":               Path("Orbital_Delivery_Schedule.csv"),
}

missing_files = [name for name, p in FILES.items() if not p.exists()]
if missing_files:
    sys.exit(f"❌ These files were not found in the working folder:\n    " +
             "\n    ".join(missing_files))

# ---------------------------------------------------------------------------
# Collect UPCs file‑by‑file
# ---------------------------------------------------------------------------
print("Scanning…  (large workbooks may take a few minutes)\n")

upc_sets: dict[str, set[str]] = {}

for label, path in FILES.items():
    try:
        if path.suffix.lower() == ".csv":
            upc_sets[label] = upcs_from_csv(path)
        else:
            upc_sets[label] = upcs_from_xlsx(path)
        print(f"  • {label:<55}  {len(upc_sets[label]):>7,} UPCs")
    except Exception as e:
        print(f"  • {label:<55}    ERROR: {e}")
        upc_sets[label] = set()

# Logical groupings
track_labels   = [l for l in FILES if l.startswith("Orbital Track Metadata")]
products_label = "Orbital Media products with territories (1).csv"
delivery_label = "Orbital_Delivery_Schedule.csv"

track_set   = set().union(*[upc_sets[l] for l in track_labels])
products_set = upc_sets[products_label]
delivery_set = upc_sets[delivery_label]

# ---------------------------------------------------------------------------
# Build presence matrix & missing‑UPC table
# ---------------------------------------------------------------------------
all_upcs = set().union(*upc_sets.values())

presence_rows = []
for u in sorted(all_upcs):
    row = {"UPC": u}
    for lbl in FILES:
        row[lbl] = u in upc_sets[lbl]
    presence_rows.append(row)

presence_df = pd.DataFrame(presence_rows)

missing_upcs = sorted((track_set | products_set) - delivery_set)
missing_df = presence_df[presence_df["UPC"].isin(missing_upcs)]

summary_df = pd.DataFrame(
    [{"Item": lbl, "Unique UPCs": len(upc_sets[lbl])} for lbl in FILES] +
    [{"Item": "UPCs present in ALL files",
      "Unique UPCs": len(set.intersection(*upc_sets.values()))}] +
    [{"Item": "UPCs in track metadata or products but missing from delivery",
      "Unique UPCs": len(missing_upcs)}]
)

# ---------------------------------------------------------------------------
# Write CSV outputs
# ---------------------------------------------------------------------------
presence_df.to_csv("orbital_upc_presence_matrix.csv", index=False)
missing_df.to_csv("orbital_upc_missing_breakdown.csv", index=False)
summary_df.to_csv("orbital_upc_summary_counts.csv", index=False)

print("\n✅  Done!  You will find three new CSV files in the folder:")
print("    • orbital_upc_presence_matrix.csv")
print("    • orbital_upc_missing_breakdown.csv")
print("    • orbital_upc_summary_counts.csv")
# ---------------------------------------------------------------------------
